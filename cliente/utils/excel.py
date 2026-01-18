from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def auto_ajustar_columnas(worksheet):
    """
    Ajusta automáticamente el ancho de las columnas
    según el contenido más largo
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column  # número de columna

        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        worksheet.column_dimensions[
            get_column_letter(column)
        ].width = max_length + 2


def exportar_leads_excel(modeladmin, request, queryset):
    """
    Exporta los leads seleccionados (o todos)
    a un archivo Excel (.xlsx)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Encabezados
    headers = [
        "Nombre",
        "Teléfono",
        "Email",
        "Ciudad",
        "Estado",
        "Tipo de financiamiento",
        "Valor del bien",
        "Entrada",
        "Plazo (meses)",
        "Ingreso mensual",
        "Tipo de empleo",
        "Tiempo empleo (meses)",
        "Fecha de creación",
    ]

    ws.append(headers)

    # Estilo encabezados
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Datos
    for lead in queryset:
        ws.append([
            lead.nombre,
            lead.telefono,
            lead.email,
            lead.ciudad,
            lead.estado,
            lead.tipo_financiamiento,
            lead.valor_bien,
            lead.entrada,
            lead.plazo_meses,
            lead.ingreso_mensual,
            lead.tipo_empleo,
            lead.tiempo_empleo_meses,
            lead.creado_en.strftime("%d/%m/%Y %H:%M") if lead.creado_en else "",
        ])

    # Opciones Excel
    auto_ajustar_columnas(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Respuesta HTTP
    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = 'attachment; filename="leads.xlsx"'

    wb.save(response)
    return response
